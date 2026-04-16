import re
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Exact colors from the template ──────────────────────────────────────────
NAVY_HEX      = "FF2F5496"   # Row 16 header fill + Row 1 font
WHITE_HEX     = "FFFFFFFF"
META_FONT_HEX = "FF666666"   # Row 2 metadata text
YELLOW_HEX    = "FFFFFF00"   # WO% column fill
GRAY_ROW_HEX  = "FFF2F2F2"  # Alternating row fill
TOTAL_FILL    = "FFDCE6F1"   # Grand totals row fill

# ── Column widths (exact from template) ─────────────────────────────────────
COL_WIDTHS = {
    "A": 6.0,     # #
    "B": 25.0,    # Section / Room
    "C": 19.33,   # Trade
    "D": 34.11,   # Description
    "E": 10.0,    # Qty
    "F": 7.33,    # Unit
    "G": 13.11,   # Unit Price
    "H": 10.0,    # Tax
    "I": 12.0,    # O&P
    "J": 14.0,    # RCV
    "K": 16.11,   # Depreciation
    "L": 14.0,    # ACV
    "M": 17.89,   # Labor
    "N": 15.89,   # Materials
    "O": 20.0,    # WO%
    "P": 20.0,    # WO Labor ONLY
    "Q": 22.66,   # WO L&M
}

HEADERS = [
    "#", "Section / Room", "Trade", "Description",
    "Qty", "Unit", "Unit Price", "Tax", "O&P", "RCV",
    "Depreciation", "ACV", "Labor", "Materials",
    "WO% ", "WO Labor ONLY", "WO L&M",
]

SOURCE_LABELS = [
    "Pulled from pdf",                          # A
    "Pulled from pdf",                          # B
    "Input manually by user",                   # C
    "Pulled from pdf",                          # D
    "Pulled from pdf",                          # E
    "Pulled from pdf",                          # F
    "Pulled from pdf",                          # G
    "Pulled from pdf",                          # H
    "Pulled from pdf",                          # I
    "Pulled from pdf",                          # J
    "Pulled from pdf - but No need to track this",  # K
    "Pulled from pdf - but No need to track this",  # L
    "Calculated here",                          # M
    "Input to set tax amount at 6%",            # N
    "This will be a percentage set per the user",   # O
    "Calculation",                              # P
    "Calculation",                              # Q
]


def _solid(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = "FF000000", bold: bool = False, size: float = 10.0) -> Font:
    return Font(color=hex_color, bold=bold, size=size, name="Arial")


def _border_bottom() -> Border:
    thin = Side(style="thin", color="FFD0D0D0")
    return Border(bottom=thin)


def build_excel(
    metadata: dict,
    line_items: list[dict],
    output_path: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate Line Items"

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    insured = metadata.get("insured_name", "Unknown Insured")
    ws["A1"] = f"{insured.upper()} - Insurance Repair Estimate"
    ws["A1"].font = _font(NAVY_HEX, bold=True, size=14.0)
    ws.row_dimensions[1].height = 22

    # ── Row 2: Metadata strip ─────────────────────────────────────────────────
    insurer   = metadata.get("insurance_company", "")
    claim     = metadata.get("claim_number", "")
    loss_type = metadata.get("loss_type", "")
    address   = metadata.get("property_address", "")
    meta_str  = f"{insurer} | Claim: {claim} | Type: {loss_type} | {address}"
    ws["A2"] = meta_str
    ws["A2"].font = _font(META_FONT_HEX, bold=False, size=9.0)
    ws.row_dimensions[2].height = 14

    # ── Rows 3–14: whitespace ─────────────────────────────────────────────────
    for r in range(3, 15):
        ws.row_dimensions[r].height = 8

    # ── Row 15: Source annotation labels ─────────────────────────────────────
    for col_idx, label in enumerate(SOURCE_LABELS, start=1):
        cell = ws.cell(row=15, column=col_idx, value=label)
        cell.font = _font("FF888888", size=8.0)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[15].height = 28

    # ── Row 16: Column headers ────────────────────────────────────────────────
    header_fill = _solid(NAVY_HEX)
    header_font = _font(WHITE_HEX, bold=True, size=11.0)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=16, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 18

    # ── Data rows (17+) ───────────────────────────────────────────────────────
    first_data_row = 17
    gray_fill = _solid(GRAY_ROW_HEX)
    wo_font_color = "FFFF0000"  # red for WO% column (template uses red)
    num_fmt_currency = '#,##0.00'
    num_fmt_pct = '0%'

    for idx, item in enumerate(line_items):
        row = first_data_row + idx
        row_fill = gray_fill if idx % 2 == 1 else None

        def w(col: int, val, number_format: str = None, font: Font = None, fill: PatternFill = None):
            c = ws.cell(row=row, column=col, value=val)
            if row_fill and fill is None:
                c.fill = row_fill
            elif fill:
                c.fill = fill
            c.font = font or _font(size=10.0)
            c.alignment = Alignment(vertical="center")
            if number_format:
                c.number_format = number_format
            return c

        w(1,  idx + 1)                                       # A: line #
        w(2,  item.get("section", ""))                       # B: Section
        w(3,  "")                                            # C: Trade (user-entered)
        w(4,  item.get("description", ""))                   # D: Description
        w(5,  _safe_float(item.get("qty")))                  # E: Qty
        w(6,  item.get("unit", ""))                          # F: Unit
        w(7,  _safe_float(item.get("unit_price")), num_fmt_currency)   # G
        w(8,  _safe_float(item.get("tax")),        num_fmt_currency)   # H
        w(9,  _safe_float(item.get("o_and_p")),    num_fmt_currency)   # I
        w(10, _safe_float(item.get("rcv")),        num_fmt_currency)   # J
        w(11, _safe_float(item.get("depreciation")), num_fmt_currency) # K
        w(12, _safe_float(item.get("acv")),        num_fmt_currency)   # L

        # M: Labor = RCV - O&P - Tax - Materials  (formula)
        mc = ws.cell(row=row, column=13, value=f"=J{row}-I{row}-H{row}-N{row}")
        if row_fill:
            mc.fill = row_fill
        mc.font = _font(size=10.0)
        mc.number_format = num_fmt_currency

        # N: Materials = Tax / 0.06  (formula)
        nc = ws.cell(row=row, column=14, value=f"=H{row}/0.06")
        if row_fill:
            nc.fill = row_fill
        nc.font = _font(size=10.0)
        nc.number_format = num_fmt_currency

        # O: WO% — yellow fill, user input
        oc = ws.cell(row=row, column=15, value=None)
        oc.fill = _solid(YELLOW_HEX)
        oc.font = _font(wo_font_color, size=10.0)
        oc.number_format = num_fmt_pct

        # P: WO Labor Only = Labor * WO%
        pc = ws.cell(row=row, column=16, value=f"=M{row}*O{row}")
        if row_fill:
            pc.fill = row_fill
        pc.font = _font(size=10.0)
        pc.number_format = num_fmt_currency

        # Q: WO L&M = (Materials + Labor) * WO%
        qc = ws.cell(row=row, column=17, value=f"=(N{row}+M{row})*O{row}")
        if row_fill:
            qc.fill = row_fill
        qc.font = _font(size=10.0)
        qc.number_format = num_fmt_currency

        ws.row_dimensions[row].height = 15

    # ── Grand totals row ──────────────────────────────────────────────────────
    last_data_row = first_data_row + len(line_items) - 1
    totals_row = last_data_row + 2
    totals_fill = _solid(TOTAL_FILL)
    totals_font = _font(NAVY_HEX, bold=True, size=10.0)

    ws.cell(row=totals_row, column=1, value="GRAND TOTALS").font = totals_font
    ws.cell(row=totals_row, column=1).fill = totals_fill

    for col in range(2, 18):
        c = ws.cell(row=totals_row, column=col)
        c.fill = totals_fill
        c.font = totals_font

    for col_idx, col_letter in enumerate(["H", "I", "J", "K", "L"], start=8):
        c = ws.cell(row=totals_row, column=col_idx)
        c.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        c.number_format = '#,##0.00'

    ws.row_dimensions[totals_row].height = 18

    # ── Column widths ────────────────────────────────────────────────────────
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze panes at A17 ──────────────────────────────────────────────────
    ws.freeze_panes = "A17"

    wb.save(output_path)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r"[$,]", "", s)
    # parentheses → negative
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0
